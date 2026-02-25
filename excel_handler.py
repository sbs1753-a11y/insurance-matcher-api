import openpyxl


def find_row_by_label(excel_path, label, sheet_name=None, search_col=2):
    """특정 텍스트가 있는 행 번호를 자동 탐지"""
    wb = openpyxl.load_workbook(excel_path)
    ws = wb[sheet_name] if sheet_name else wb.active
    for row_idx in range(1, ws.max_row + 1):
        val = ws.cell(row=row_idx, column=search_col).value
        if val and label in str(val).strip():
            wb.close()
            return row_idx
    wb.close()
    return None


def find_insurer_row(excel_path, sheet_name=None, search_col=2):
    """보험사명이 들어갈 행 자동 탐지"""
    wb = openpyxl.load_workbook(excel_path)
    ws = wb[sheet_name] if sheet_name else wb.active
    for row_idx in range(1, ws.max_row + 1):
        val = ws.cell(row=row_idx, column=search_col).value
        if val and "회사" in str(val).strip():
            wb.close()
            return row_idx + 1
    wb.close()
    return None


def find_structure(excel_path, sheet_name=None, search_col=2):
    """엑셀 보장분석표 구조 자동 탐지"""
    wb = openpyxl.load_workbook(excel_path)
    ws = wb[sheet_name] if sheet_name else wb.active

    structure = {
        "insurer_row": None,
        "product_row": None,
        "premium_row": None,
        "reserve_row": None,
        "start_row": None,
    }

    for row_idx in range(1, ws.max_row + 1):
        val = ws.cell(row=row_idx, column=search_col).value
        if val is None:
            continue
        val_str = str(val).strip()

        if "회사" in val_str and structure["insurer_row"] is None:
            structure["insurer_row"] = row_idx + 1
        if "상품" in val_str and structure["product_row"] is None:
            structure["product_row"] = row_idx
        if "보험료" in val_str and "총" not in val_str and structure["premium_row"] is None:
            structure["premium_row"] = row_idx
        if "적립금" in val_str and structure["reserve_row"] is None:
            structure["reserve_row"] = row_idx
        if val_str in ["실비질병/상해 종합입원", "실비질병/상해종합입원"] and structure["start_row"] is None:
            structure["start_row"] = row_idx

    if structure["start_row"] is None and structure["premium_row"]:
        structure["start_row"] = structure["premium_row"] + 3

    wb.close()
    return structure


def read_excel_coverages(excel_path, sheet_name=None, coverage_col=2, amount_col=4, start_row=8):
    """Excel 보장분석표에서 특약명 목록 읽기"""
    wb = openpyxl.load_workbook(excel_path)
    ws = wb[sheet_name] if sheet_name else wb.active

    coverages = []
    skip_values = [
        "주계약", "특약", "합계", "총보험료", "보장항목",
        "담보명", "특약명", ""
    ]

    for row_idx in range(start_row, ws.max_row + 1):
        cell_value = ws.cell(row=row_idx, column=coverage_col).value

        if cell_value is None:
            continue

        name = str(cell_value).strip()

        if name in skip_values:
            continue

        if len(name) < 2:
            continue

        coverages.append({
            "row": row_idx,
            "특약명": name,
            "amount_col": amount_col
        })

    wb.close()
    return coverages


def write_matched_amounts(excel_path, output_path, matched_data, sheet_name=None):
    """매칭 결과를 Excel에 기록"""
    wb = openpyxl.load_workbook(excel_path)
    ws = wb[sheet_name] if sheet_name else wb.active

    for item in matched_data:
        ws.cell(
            row=item["row"],
            column=item["amount_col"],
            value=item["가입금액"]
        )

    wb.save(output_path)
    wb.close()


def write_insurer_info(excel_path, output_path, insurer_name, insurer_row, product_name, product_row, col, sheet_name=None):
    """보험사명과 상품명을 Excel 특정 셀에 기록"""
    wb = openpyxl.load_workbook(excel_path)
    ws = wb[sheet_name] if sheet_name else wb.active

    ws.cell(row=insurer_row, column=col, value=insurer_name)
    ws.cell(row=product_row, column=col, value=product_name)

    wb.save(output_path)
    wb.close()


def write_premium(excel_path, output_path, premium, premium_row, col, sheet_name=None):
    """보험료를 Excel 특정 셀에 기록"""
    wb = openpyxl.load_workbook(excel_path)
    ws = wb[sheet_name] if sheet_name else wb.active

    ws.cell(row=premium_row, column=col, value=premium)

    wb.save(output_path)
    wb.close()
